import requests
import re
import datetime
from openpyxl import Workbook
import json
from jsonpath import jsonpath
import openpyxl


# 用户名密码方式
# def get_tunnel_proxy():
#     # 隧道域名:端口号
#     "tps764.kdlapi.com:15818"
#     # 用户名密码方式
#     username = "t14562671881861"
#     password = "7hjgwxtr"
#     proxies = {
#         "http": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": tunnel},
#         "https": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": tunnel}
#     }
#     return proxies

# 隧道域名:端口号
tunnel = "tps764.kdlapi.com:15818"

# 用户名密码方式
username = "t14562671881861"
password = "7hjgwxtr"
proxies = {
    "http": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": tunnel},
    "https": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": tunnel}
}

# 设置一个计数器
con = 0
# 伪装浏览器抬头
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36',
    'connect': 'close'
}
# 打开 ADD2.txt文件 进行读取 BTC地址
with open('C:/Users/25472/Desktop/23000.txt', 'r')as f:
    # 读取每一行
    for line in f:
        # 每读取一行 计数器+1
        con += 1

        # 设置列表，进行保存抓取下来的数据
        TIME = []
        ONEDAY = []
        INPUTSVALUE = []
        OUTSVALUE = []
        ADDRESINPUTS = []
        ADDRESOUTS = []
        ONEDAYCHA = []

        # TXT文件数据清洗，去掉所有回车空格等
        #lines = line.strip()
        lines = line

        # 设置未知次数循环，最大99999 ，类似于while，满足条件停止循环
        for i in range(0, 99999, 100):
            # 地址，i为页数，因为不知道具体有多少页，不停的进行翻页
            url = 'https://blockchain.info/rawaddr/%s?limit=100&offset=%s' % (lines, i)
            print(url)
            print(i, '我是页数')


            # 设置一个刷新页面计数器
            max_retry = 0
            # 设置最多刷新三次，成功返回，页面卡住的话 ， 刷新重新打开API
            while max_retry < 3:
                try:
                    res = requests.get(url=url, headers=headers, proxies=proxies, timeout=10).text
                    jsontext = json.loads(res)
                    break
                except Exception:
                    print('网页卡住了,重新运行')
                max_retry += 1

  # 这是设置了一个末页提醒，如果抓到的页面<500个字节，说明数据已经抓完了，并且此页是没有数据的，直接break就行了
            if len(res) < 500:
                break

 # jsonpath 包 ，获取到交易时间
            time = jsonpath(jsontext, '$..time')

 # 将交易时间 ，追加近大写的TIME中，保存
            TIME.extend(time)


# 每一次的所有交易==================================================
            for i1 in range(0, len(time), 1):
                try:
                    inputs = len(jsonpath(jsontext, 'txs[%s]..inputs..addr' % i1))
                    outs = len(jsonpath(jsontext, 'txs[%s]..out..addr' % i1))
                except:
                    continue
                AA = inputs + outs
                BB = inputs - outs
                ONEDAYCHA.append(BB)
                ONEDAY.append(AA)

            # 收入价格
            inputsall = jsonpath(jsontext, '$..inputs..value')
            # 出去价格
            outsall = jsonpath(jsontext, '$..out..value')
            INPUTSVALUE.extend(inputsall)
            OUTSVALUE.extend(outsall)

            # 抓取所有的地址分为  收入地址(inputs)   和    支出地址（outs）
            addrinputs = jsonpath(jsontext, '$..inputs..addr')
            ADDRESINPUTS.extend(addrinputs)
            addrouts = jsonpath(jsontext, '$..out..addr')
            ADDRESOUTS.extend(addrouts)

            if len(time) < 100:
                break

        ind1 = []
        ind2 = []
        find = lines

        # 循环inputs（收入BTC）的数据
        for i6, v in enumerate(ADDRESINPUTS):
            if v == find:
                ind1.append(i6)

        # 循环outs（支出BTC）的数据
        for i7, v in enumerate(ADDRESOUTS):
            if v == find:
                ind2.append(i7)
        # 设置两个列表，保存数据
        ind1sum = []
        ind2sum = []
        # num1-16 就是最终出来的数据结果 顺序是按照你上次给我的要求，一条一条排下去的
        #         第一个 时间差
        try:
            num1 = int(TIME[0]) - int(TIME[-1])

        except:
            print('垃圾地址，注册了不交易')
            continue
        # 第二个 最大交易量  以此类推
        num2 = max(ONEDAY)
        num3 = len(TIME)
        num4 = len(INPUTSVALUE) / len(OUTSVALUE)
        num5 = len(ind1)
        num6 = len(ind2)
        num7_1 = len(ADDRESINPUTS)
        num7_2 = len(ADDRESOUTS)
        YANCHI = []
        for i3 in range(0, len(TIME) - 1):
            yanchi = int(TIME[i3]) - int(TIME[i3 + 1])
            YANCHI.append(yanchi)
        YANCHI.sort()

        # 计算中位数
        half = len(YANCHI) // 2
        SSS = (YANCHI[half] + YANCHI[~half]) / 2
        num8 = int(SSS)
        num9 = sum(YANCHI) / len(YANCHI)
        num10 = min(YANCHI)
        num11 = max(YANCHI)
        for i9 in ind1:
            ind1sum.append(INPUTSVALUE[i9])
        for i9 in ind2:
            ind2sum.append(OUTSVALUE[i9])

        # API中的 价格/100000000 为页面中显示的价格
        num12 = sum(ind1sum) / 100000000
        num13 = sum(ind2sum) / 100000000

        num14 = (num12) / (len(ind1sum)+0.0001)
        num15 = (num13) / (len(ind2sum)+0.0001)
        TIME48 = []

        # 计算时间差
        cha = 0
        for i4 in range(0, len(TIME) - 1):
            # 172800为一天的时间戳
            if int(TIME[i4]) + 172800 <= int(TIME[i4 + 1]):
                cha = ONEDAYCHA[i4] + ONEDAYCHA[i4 + 1] + cha
            else:
                if cha != 0:
                    TIME48.append(int(cha))
                else:
                    TIME48.append(int(ONEDAYCHA[i4]))
                    cha = 0
        TIME49 = []
        for i5 in TIME48:
            if i5 <= 0:
                AAA = abs(i5)
                TIME49.append(AAA)
            else:
                TIME49.append(i5)
        maxnum = max(TIME49)
        TIMEEND = TIME49.index(maxnum)
        num16 = TIME48[TIMEEND]

        # 打印抓到的所有数据，顺序是按照你上次给我的要求，一条一条排下去的
        print(num1)
        print(num2)
        print(num3)
        print(num4)
        print(num5)
        print(num6)
        print(num7_1)
        print(num7_2)
        print(num8)
        print(num9)
        print(num10)
        print(num11)
        print(num12)
        print(num13)
        print(num14)
        print(num15)
        print(num16)

        # 打开ADDRS.XLSX的表格
        wb = openpyxl.load_workbook('C:/Users/25472/Desktop/addrs-25.xlsx')
        # 打开第一个sheet
        ws1 = wb.worksheets[0]
        # 将所有的数据 依次写入addrs.xlsx的表格中

        ws1.cell(row=con, column=2, value=num1)
        ws1.cell(row=con, column=3, value=num2)
        ws1.cell(row=con, column=4, value=num3)
        ws1.cell(row=con, column=5, value=num4)
        ws1.cell(row=con, column=6, value=num5)
        ws1.cell(row=con, column=7, value=num6)
        ws1.cell(row=con, column=8, value=num7_1)
        ws1.cell(row=con, column=9, value=num7_2)
        ws1.cell(row=con, column=10, value=num8)
        ws1.cell(row=con, column=11, value=num9)
        ws1.cell(row=con, column=12, value=num10)
        ws1.cell(row=con, column=13, value=num11)
        ws1.cell(row=con, column=14, value=num12)
        ws1.cell(row=con, column=15, value=num13)
        ws1.cell(row=con, column=16, value=num14)
        ws1.cell(row=con, column=17, value=num15)
        ws1.cell(row=con, column=18, value=num16)
        # 表格保存
        wb.save('C:/Users/25472/Desktop/addrs-25.xlsx')
        # 循环继续
        print('下一个地址')
        print(con)


print ("退出主线程")